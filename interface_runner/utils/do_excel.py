from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import BLACK
from collections import namedtuple

import os


class DoExcel:

    def __init__(self, file):
        """
        :param dir: excel文件名，file要包括路径
        """
        try:
            self.file_name = file

            # current_path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
            self.__wb = load_workbook(file)  # 打开excel
        except FileNotFoundError as e:

            raise e

    def get_sheet_name(self):
        """
        :description:获取对应excel文件所有sheet名，返回list
        :return: sheetNames = []
        """
        sheetNames = self.__wb.sheetnames
        return sheetNames

    def excel_close(self):
        self.__wb.close()

    def get_max_row_num(self, sheet_name):
        """获取最大行号"""
        max_row_num = self.__wb[sheet_name].max_row
        return max_row_num

    def get_max_column_num(self, sheet_name):
        """获取最大列号"""
        max_column = self.__wb[sheet_name].max_column
        return max_column

    def get_cell_value(self, sheet_name, coordinate=None, row=None, column=None):
        """获取指定单元格的数据"""
        if coordinate is not None:
            try:
                return self.__wb[sheet_name][coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and row is not None and column is not None:
            if isinstance(row, int) and isinstance(column, int):
                return self.__wb[sheet_name].cell(row=row, column=column).value
            else:
                raise TypeError('row and column must be type int')
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def get_row_value(self, sheet_name, row):
        """获取某一行的数据"""
        column_num = self.get_max_column_num(sheet_name)
        row_value = []
        if isinstance(row, int):
            for column in range(1, column_num + 1):
                values_row = self.__wb[sheet_name].cell(row, column).value
                row_value.append(values_row)
            return row_value
        else:
            raise TypeError('row must be type int')

    def get_column_value(self, sheet_name, column_name):
        """获取某一列数据"""
        index = self.get_excel_title(sheet_name).index(column_name) + 1
        row_num = self.get_max_row_num(sheet_name)
        column_value = []
        if isinstance(index, int):
            for row in range(1, row_num + 1):
                values_column = self.__wb[sheet_name].cell(row, index).value
                column_value.append(values_column)
            return column_value
        else:
            raise TypeError('column must be type int')

    def get_all_value_1(self, sheet_name):
        """获取指定表单的所有数据(除去表头)"""
        max_row_num = self.get_max_row_num(sheet_name)

        max_column = self.get_max_column_num(sheet_name)
        sheet = self.__wb[sheet_name]
        values = []
        for row in range(2, max_row_num + 1):
            value_list = []
            for column in range(1, max_column + 1):
                value = sheet.cell(row, column).value
                value_list.append(value)
            values.append(value_list)
        return values

    def get_all_value_2(self, sheet_name):
        """获取指定表单的所有数据(除去表头)"""

        rows_obj = self.__wb[sheet_name].iter_rows(min_row=2, max_row=self.__wb[sheet_name].max_row, values_only=True)
        values = []
        for row_tuple in rows_obj:
            value_list = []
            n = 0
            for value in row_tuple:
                if value == None:  # 添加逻辑，若获取到的表格value为None，则重新赋值为字符串
                    value = ""
                    n = n + 1
                value_list.append(value)
            if n >= 12:  # 若一行有12个及以上的None，则不返回数据（删除空行数据）
                pass
            else:
                values.append(value_list)
        return values

    def get_excel_title(self, sheet_name):
        """获取sheet表头"""
        title_key = tuple(self.__wb[sheet_name].iter_rows(max_row=1, values_only=True))[0]
        return title_key

    def get_listdict_all_value(self, sheet_name):
        """获取所有数据，返回嵌套字典的列表"""
        sheet_title = self.get_excel_title(sheet_name)
        all_values = self.get_all_value_2(sheet_name)
        value_list = []
        for value in all_values:
            value_list.append(dict(zip(sheet_title, value)))
        return value_list

    def get_row_values(self, sheet_name):
        """获取所有行的数据，返回字典，第一行为key"""
        value_dict = {}
        row_num = self.get_max_row_num(sheet_name)

        for i in range(1, row_num + 1):
            value = self.get_row_value(sheet_name, i)
            value_dict[value[0]] = value[1:]
        return value_dict

    def write_cell(self, sheet_name, row, column, value=None, bold=True, color=BLACK):
        if isinstance(row, int) and isinstance(column, int):
            try:
                cell_obj = self.__wb[sheet_name].cell(row, column)
                cell_obj.font = Font(color=color, bold=bold)
                cell_obj.value = value
                self.__wb.save(self.file_name)
                self.__wb.close()
            except Exception as e:
                raise e
        else:
            raise TypeError('row and column must be type int')


if __name__ == '__main__':
    ex = DoExcel(r'C:\Users\060208191\Desktop\member_case_b.xlsx')
    # out=ex.get_cell_value_by_row("pc_xpath")
    out = ex.get_listdict_all_value("my")
    sheetNames = ex.get_sheet_name()

    print(out)

    # c
